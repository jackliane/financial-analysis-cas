import streamlit as st
import openpyxl
import pandas as pd
from io import BytesIO

st.set_page_config(page_title="财务数据分析", layout="wide")
st.title("🎈 该网站主要用于对财务数据进行分析")

uploaded_file = st.file_uploader(
    '上传 Excel 文件，其中请确保空壳公司sheet命名为"空壳"，注销公司sheet命名为"注销"，需要处理的数据sheet命名为"需要处理"，确保需要处理的数据sheet中单位名称列命名为“账户名”',
    type=["xlsx"],
    help='支持 .xlsx 格式的文件'
)

if uploaded_file is not None:
    # 1. 直接从上传的文件对象加载（不需要路径）
    wb = openpyxl.load_workbook(uploaded_file)

    if '需要处理' not in wb.sheetnames:
        st.error('❌ 未找到名为"需要处理"的sheet，请检查上传文件')
    else:
        sheet = wb['需要处理']

        # 填充空白的公司名称（向下填充）
        name_previous = ''
        for row in range(2, sheet.max_row + 1):
            name = str(sheet[f'A{row}'].value)
            if name == 'None':
                name = name_previous
                sheet[f'A{row}'] = name
            else:
                name_previous = name

        # 2. 将处理后的主数据保存到内存（BytesIO），而非磁盘
        output_main = BytesIO()
        wb.save(output_main)
        output_main.seek(0)  # 重置指针到开头

        # 3. 从内存读取数据
        data = pd.read_excel(output_main, sheet_name='需要处理')

        # 读取空壳和注销列表
        kongke_list = pd.read_excel(output_main, sheet_name='空壳')['原文件导入名称'].unique().tolist()
        output_main.seek(0)
        zhuxiao_list = pd.read_excel(output_main, sheet_name='注销')['原文件导入名称'].unique().tolist()
        output_main.seek(0)

        # 筛选数据
        data_kongke = data.query('账户名 in @kongke_list')
        data_zhuxiao = data.query('账户名 in @zhuxiao_list')

        # ---- 处理空壳数据：合并相邻重名 ----
        kongke_buf = BytesIO()
        data_kongke.to_excel(kongke_buf, index=False)
        kongke_buf.seek(0)

        wb = openpyxl.load_workbook(kongke_buf)
        sheet = wb['Sheet1']

        for row in range(sheet.max_row, 1, -1):
            name = str(sheet[f'A{row}'].value)
            name_previous = str(sheet[f'A{row - 1}'].value)
            if name == name_previous:
                sheet[f'A{row}'] = ''

        kongke_output = BytesIO()
        wb.save(kongke_output)
        kongke_output.seek(0)
        wb.close()

        # ---- 处理注销数据：合并相邻重名 ----
        zhuxiao_buf = BytesIO()
        data_zhuxiao.to_excel(zhuxiao_buf, index=False)
        zhuxiao_buf.seek(0)

        wb = openpyxl.load_workbook(zhuxiao_buf)
        sheet = wb['Sheet1']

        for row in range(sheet.max_row, 1, -1):
            name = str(sheet[f'A{row}'].value)
            name_previous = str(sheet[f'A{row - 1}'].value)
            if name == name_previous:
                sheet[f'A{row}'] = ''

        zhuxiao_output = BytesIO()
        wb.save(zhuxiao_output)
        zhuxiao_output.seek(0)
        wb.close()

        # 准备主数据文件供下载
        output_main.seek(0)
        wb.close()

        # ---- 展示结果 ----
        st.success("✅ 数据处理完成！")

        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("总记录数", len(data))
        with col2:
            st.metric("空壳公司记录", len(data_kongke))
        with col3:
            st.metric("注销公司记录", len(data_zhuxiao))

        st.divider()

        # 4. 通过 download_button 让用户下载处理后的文件
        st.subheader("📥 下载处理结果")

        dl_col2, dl_col3 = st.columns(2)

        with dl_col2:
            st.download_button(
                label="📄 下载空壳公司数据",
                data=kongke_output,
                file_name="空壳.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        with dl_col3:
            st.download_button(
                label="📄 下载注销公司数据",
                data=zhuxiao_output,
                file_name="注销.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # 预览数据
        with st.expander("🔍 预览空壳公司数据"):
            st.dataframe(data_kongke.head(50), use_container_width=True)

        with st.expander("🔍 预览注销公司数据"):
            st.dataframe(data_zhuxiao.head(50), use_container_width=True)

else:
    st.info("👆 请上传一个 .xlsx 文件")
